"""
检查Excel文件内容
"""
from openpyxl import load_workbook

excel_path = r'f:\trace\work-space\yys-show\式神表.xlsx'
wb = load_workbook(excel_path)
ws = wb.active

print(f'工作表名称: {ws.title}')
print(f'总行数: {ws.max_row}')
print(f'总列数: {ws.max_column}')
print()

print('前10行原始数据:')
for i in range(1, min(11, ws.max_row + 1)):
    row_data = []
    for j in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=i, column=j).value
        row_data.append(cell_value)
    print(f'  第{i}行: {row_data}')
