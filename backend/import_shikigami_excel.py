"""
从Excel导入式神数据到数据库
"""
import sys
import os

# 添加项目根目录到Python路径
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from app import create_app
from app.models.shikigami import Shikigami
from app.models import db


def import_from_excel():
    """从Excel导入式神数据"""
    app = create_app()
    with app.app_context():
        # 读取Excel
        excel_path = r'f:\trace\work-space\yys-show\式神表.xlsx'
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # 获取表头
        headers = [cell.value for cell in ws[1]]
        print('Excel列名:', headers)
        print(f'\n总共 {ws.max_row - 1} 行数据')
        print('\n前5行数据:')
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), 2):
            print(f'  第{i}行: {row}')
        
        # 统计
        success_count = 0
        skip_count = 0
        error_count = 0
        
        # 从第2行开始读取（跳过表头）
        for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            try:
                # 获取式神名称（假设第一列是名称）
                name = str(row[0]).strip() if row[0] else None
                
                if not name or name == 'None' or name == 'nan':
                    print(f'跳过第 {index} 行：名称为空')
                    skip_count += 1
                    continue
                
                # 检查是否已存在
                existing = Shikigami.query.filter_by(name=name).first()
                if existing:
                    print(f'跳过：式神 "{name}" 已存在')
                    skip_count += 1
                    continue
                
                # 创建式神
                shikigami = Shikigami(
                    name=name,
                    english_name='',
                    rarity='SR'  # 默认SR
                )
                
                db.session.add(shikigami)
                success_count += 1
                print(f'准备添加: {name}')
                
                # 每10条提交一次
                if success_count % 10 == 0:
                    db.session.commit()
                    print(f'>>> 已提交 {success_count} 条...')
                
            except Exception as e:
                print(f'第 {index} 行导入失败: {e}')
                error_count += 1
                db.session.rollback()
        
        # 最后提交
        if success_count > 0:
            db.session.commit()
            print(f'>>> 最终提交完成')
        
        print(f'\n导入完成！')
        print(f'成功: {success_count}')
        print(f'跳过: {skip_count}')
        print(f'失败: {error_count}')
        
        # 验证导入结果
        total = Shikigami.query.count()
        print(f'\n数据库中现有 {total} 个式神')


if __name__ == '__main__':
    import_from_excel()
